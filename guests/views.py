from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.models import User
from django.views.decorators.http import require_POST
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse, Http404, HttpResponseForbidden
from .models import GuestEntry, FollowUpReport, SocialMediaEntry, Review
from .forms import GuestEntryForm, FollowUpReportForm
import csv
import io
from django.utils.dateparse import parse_date
from django.contrib.auth import get_user_model, authenticate, login
from django.core.paginator import Paginator
from django.contrib.auth.models import Group, User
from django.db.models import Q, Count, Max, F, Prefetch
from django.utils.http import urlencode
import openpyxl
from openpyxl.utils import get_column_letter
from django.contrib import messages
from datetime import datetime, timedelta
from django.utils import timezone
from django.utils.timezone import localtime, now, localdate, is_naive, make_aware
from django.utils.timesince import timesince
import pytz
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from django.template.loader import render_to_string, get_template
#import weasyprint
#from .utils import get_week_start_end
from django.utils.dateparse import parse_date
from django.db.models.functions import ExtractYear, ExtractMonth, TruncMonth
import calendar
import base64
import json
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
import os
from django.db import IntegrityError, transaction
from django.middleware.csrf import get_token
from urllib.parse import urlencode
from django.conf import settings
from cloudinary.uploader import upload as cloudinary_upload
from accounts.models import CustomUser, TeamMembership
from workforce.models import Event, AttendanceRecord, PersonalReminder, Team
from urllib.parse import urlencode
from accounts.utils import (
    user_in_groups,
    is_project_wide_admin,
    is_magnet_admin,
    is_team_admin,
    is_project_admin,
    is_project_level_role,
    user_in_team,
)
from workforce.consumers import get_user_color
from workforce.utils import get_available_events_for_user, get_calendar_items, expand_team_events, get_visible_clock_records



User = get_user_model()


@login_required
def dashboard_view(request):
    """Optimized user dashboard view with all stats, team data, and events."""
    user = request.user
    today = timezone.localdate()
    current_year = today.year
    last_30_days = today - timedelta(days=30)

    # ---------------------- Base queryset ----------------------
    if is_magnet_admin(user):
        queryset = GuestEntry.objects.all()
    else:
        queryset = GuestEntry.objects.filter(assigned_to=user)

    # ---------------------- Available years ----------------------
    guest_entries = GuestEntry.objects.all()
    available_years = [d.year for d in guest_entries.dates('date_of_visit', 'year')]

    # ---------------------- Summary JSON ----------------------
    request.GET = request.GET.copy()
    request.GET['year'] = str(current_year)
    summary_json = guest_entry_summary(request)
    summary_data = json.loads(summary_json.content.decode())

    # ---------------------- Aggregate services, purposes, status ----------------------
    agg_qs = queryset.values('service_attended', 'purpose_of_visit', 'status', 'channel_of_visit').annotate(count=Count('id'))

    # Service data
    service_data = {}
    total_services = 0
    for entry in agg_qs:
        service = entry['service_attended'] or "Not Specified"
        service_data[service] = service_data.get(service, 0) + entry['count']
        total_services += entry['count']

    service_labels = list(service_data.keys())
    service_counts = list(service_data.values())
    if service_counts:
        most_attended_service = service_labels[service_counts.index(max(service_counts))]
        most_attended_count = max(service_counts)
        attendance_rate = round((most_attended_count / total_services) * 100, 1) if total_services else 0
    else:
        most_attended_service, most_attended_count, attendance_rate = "No Data", 0, 0

    # Purpose stats
    purposes = ["Home Church", "Occasional Visit", "One-Time Visit", "Special Programme"]
    total_purposes = sum(entry['count'] for entry in agg_qs)
    purpose_stats = {
        p: {
            'count': sum(entry['count'] for entry in agg_qs if (entry['purpose_of_visit'] or "").lower() == p.lower()),
            'percentage': round((sum(entry['count'] for entry in agg_qs if (entry['purpose_of_visit'] or "").lower() == p.lower()) / total_purposes)*100,1) if total_purposes else 0
        }
        for p in purposes
    }

    home_church_count = purpose_stats["Home Church"]["count"]
    home_church_percentage = purpose_stats["Home Church"]["percentage"]
    occasional_visit_count = purpose_stats["Occasional Visit"]["count"]
    occasional_visit_percentage = purpose_stats["Occasional Visit"]["percentage"]
    one_time_visit_count = purpose_stats["One-Time Visit"]["count"]
    one_time_visit_percentage = purpose_stats["One-Time Visit"]["percentage"]
    special_programme_count = purpose_stats["Special Programme"]["count"]
    special_programme_percentage = purpose_stats["Special Programme"]["percentage"]

    # Status data
    status_data = {}
    for entry in agg_qs:
        status = entry['status'] or "Unknown"
        status_data[status] = status_data.get(status, 0) + entry['count']
    status_labels = list(status_data.keys())
    status_counts = list(status_data.values())

    # Channel progress
    channel_data = {}
    total_channels = 0
    for entry in agg_qs:
        channel = entry['channel_of_visit'] or "Unknown"
        channel_data[channel] = channel_data.get(channel, 0) + entry['count']
        total_channels += entry['count']
    channel_progress = [
        {'label': c, 'count': channel_data[c], 'percent': round((channel_data[c]/total_channels)*100,2) if total_channels else 0}
        for c in channel_data
    ]

    # Planted & other statuses
    planted_count = status_data.get("Planted", 0)
    planted_elsewhere_count = status_data.get("Planted Elsewhere", 0)
    relocated_count = status_data.get("Relocated", 0)
    work_in_progress_count = status_data.get("Work in Progress", 0)

    # ---------------------- Monthly & planted growth metrics in ONE query ----------------------
    first_day_this_month = today.replace(day=1)
    last_month_end = first_day_this_month - timedelta(days=1)
    first_day_last_month = last_month_end.replace(day=1)

    metrics = queryset.aggregate(
        total_guests=Count('id'),
        current_month_count=Count('id', filter=Q(date_of_visit__gte=first_day_this_month, date_of_visit__lte=today)),
        last_month_count=Count('id', filter=Q(date_of_visit__gte=first_day_last_month, date_of_visit__lte=last_month_end)),
        user_total_guest_entries=Count('id', filter=Q(assigned_to=user)),
        user_current_month_count=Count('id', filter=Q(assigned_to=user, date_of_visit__gte=first_day_this_month, date_of_visit__lte=today)),
        user_last_month_count=Count('id', filter=Q(assigned_to=user, date_of_visit__gte=first_day_last_month, date_of_visit__lte=last_month_end)),
        user_planted_total=Count('id', filter=Q(assigned_to=user, status="Planted")),
        user_planted_current_month=Count('id', filter=Q(assigned_to=user, status="Planted", date_of_visit__gte=first_day_this_month, date_of_visit__lte=today)),
        user_planted_last_month=Count('id', filter=Q(assigned_to=user, status="Planted", date_of_visit__gte=first_day_last_month, date_of_visit__lte=last_month_end)),
    )

    # Global increase rate
    if metrics['last_month_count'] == 0:
        increase_rate = percent_change = 100 if metrics['current_month_count'] else 0
    else:
        increase_rate = percent_change = round(((metrics['current_month_count'] - metrics['last_month_count']) / metrics['last_month_count'])*100,1)

    # User metrics
    if metrics['user_last_month_count'] == 0:
        user_diff_percent = 100 if metrics['user_current_month_count'] else 0
        user_diff_positive = True
    else:
        diff = ((metrics['user_current_month_count'] - metrics['user_last_month_count']) / metrics['user_last_month_count'])*100
        user_diff_percent = round(abs(diff),1)
        user_diff_positive = diff >= 0

    planted_growth_rate = round((metrics['user_planted_total'] / metrics['user_total_guest_entries'])*100,1) if metrics['user_total_guest_entries'] else 0

    if metrics['user_planted_last_month'] == 0:
        planted_growth_change = 100 if metrics['user_planted_current_month'] else 0
    else:
        planted_growth_change = round(((metrics['user_planted_current_month'] - metrics['user_planted_last_month']) / metrics['user_planted_last_month'])*100,1)

    # ---------------------- Teams + Other Users (optimized) ----------------------
    user_teams = Team.objects.prefetch_related(
        Prefetch(
            "memberships__user",
            queryset=CustomUser.objects.exclude(id=user.id).distinct()
        )
    ).filter(memberships__user=user).distinct()

    # All other users in these teams, excluding project admins
    other_users_qs = CustomUser.objects.filter(
        team_memberships__team__in=user_teams
    ).exclude(id=user.id).distinct()
    other_users = [u for u in other_users_qs if not is_project_admin(u)]
    for u in other_users:
        u.color = get_user_color(u.id)

    # Precompute (team, members) pairs
    team_member_pairs = [(team, [u for u in other_users if u.team_memberships.filter(team=team).exists()]) for team in user_teams]

    # ---------------------- Calendar & Clock Records ----------------------
    calendar_items = get_calendar_items(user)
    clock_records = get_visible_clock_records(user, since_date=last_30_days)
    total_clock_in = clock_records.filter(clock_in__isnull=False).count()
    total_clock_out = clock_records.filter(clock_out__isnull=False).count()
    today_record = clock_records.filter(date=today).first()

    # ---------------------- Upcoming Events ----------------------
    events = []
    for team in [None] + list(Team.objects.all()):
        events.extend(expand_team_events(user, getattr(team, "id", None)))
    upcoming_events = sorted([e for e in events if e["date"] >= today], key=lambda e: (e["date"], e.get("time","")))

    start_of_sunday = today - timedelta(days=(today.isoweekday() % 7))
    end_of_week = start_of_sunday + timedelta(days=6)
    weekly_events = [
        e for e in upcoming_events
        if start_of_sunday <= e["date"] <= end_of_week and (e.get("team_id") in [t.id for t in user_teams] or e.get("team_id") is None)
    ]
    for e in weekly_events:
        e["datetime_iso"] = f"{e['date']}T{e.get('time','00:00:00')}"
    next_events = weekly_events

    # ---------------------- Context ----------------------
    context = {
        'show_filters': False,
        'available_years': available_years,
        'current_year': current_year,
        'summary_data': summary_data,
        "service_labels": service_labels,
        "service_counts": service_counts,
        "status_labels": status_labels,
        "status_counts": status_counts,
        "channel_progress": channel_progress,
        "planted_count": planted_count,
        "planted_elsewhere_count": planted_elsewhere_count,
        "relocated_count": relocated_count,
        "work_in_progress_count": work_in_progress_count,
        "total_guests": metrics['total_guests'],
        "increase_rate": increase_rate,
        "percent_change": percent_change,
        "user_total_guest_entries": metrics['user_total_guest_entries'],
        "user_guest_entry_diff_percent": user_diff_percent,
        "user_guest_entry_diff_positive": user_diff_positive,
        "user_planted_total": metrics['user_planted_total'],
        "planted_growth_rate": planted_growth_rate,
        "planted_growth_change": planted_growth_change,
        "most_attended_service": most_attended_service,
        "most_attended_count": most_attended_count,
        "attendance_rate": attendance_rate,
        "home_church_count": home_church_count,
        "home_church_percentage": home_church_percentage,
        "occasional_visit_count": occasional_visit_count,
        "occasional_visit_percentage": occasional_visit_percentage,
        "one_time_visit_count": one_time_visit_count,
        "one_time_visit_percentage": one_time_visit_percentage,
        "special_programme_count": special_programme_count,
        "special_programme_percentage": special_programme_percentage,
        "other_users": other_users,
        "user_teams": user_teams,
        "team_member_pairs": team_member_pairs,
        "calendar_items": calendar_items,
        'total_clock_in': total_clock_in,
        'total_clock_out': total_clock_out,
        'today_clock_in': getattr(today_record,'clock_in',None),
        'today_clock_out': getattr(today_record,'clock_out',None),
        "available_events": upcoming_events,
        "next_events_for_week": next_events,
        "page_title": "Dashboard"
    }

    return render(request, "guests/dashboard.html", context)





def guest_entry_summary(request):
    year = request.GET.get('year')
    try:
        year = int(year)
    except (TypeError, ValueError):
        return JsonResponse({'error': 'Invalid year'}, status=400)

    guests = GuestEntry.objects.filter(date_of_visit__year=year)

    # Total guests for the year
    total_count = guests.count()

    # Group by month and count
    month_counts = (
        guests.annotate(month=ExtractMonth('date_of_visit'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )

    # Build dict: {1: 5, 2: 10, ..., 12: 0}
    counts_dict = {month: 0 for month in range(1, 13)}
    for entry in month_counts:
        counts_dict[entry['month']] = entry['count']

    max_count = max(counts_dict.values())
    min_count = min(counts_dict.values())
    avg_count = sum(counts_dict.values()) // 12 if counts_dict else 0

    # Find month names
    max_months = [calendar.month_name[m] for m, c in counts_dict.items() if c == max_count]
    min_months = [calendar.month_name[m] for m, c in counts_dict.items() if c == min_count]

    # Just use the first if tie
    max_month = max_months[0] if max_months else "N/A"
    min_month = min_months[0] if min_months else "N/A"

    def percent(count):
        return round((count / max_count) * 100, 1) if max_count > 0 else 0

    data = {
        'max_month': max_month,
        'max_count': max_count,
        'max_percent': percent(max_count),
        'min_month': min_month,
        'min_count': min_count,
        'min_percent': percent(min_count),
        'avg_count': avg_count,
        'avg_percent': percent(avg_count),
        'total_count': total_count,
    }

    return JsonResponse(data)


def top_services_data(request):
    top_services = (
        GuestEntry.objects
        .values('service_attended')
        .annotate(count=Count('id'))
        .order_by('-count')[:10]
    )

    # Prepare data: list of dicts with service, count
    data = list(top_services)

    # Calculate total count of these top 10 to calculate % widths
    total = sum(item['count'] for item in data) or 1  # avoid division by zero

    # Add percentage to each
    for item in data:
        item['percent'] = round((item['count'] / total) * 100, 1)

    #return JsonResponse({"disabled": True})
    return JsonResponse({'services': data})




@login_required
def services_attended_chart(request):
    """AJAX endpoint for services attended chart."""

    queryset = GuestEntry.objects.all()  # No filtering by user

    qs = queryset.values('service_attended').annotate(count=Count('id')).order_by('-count')
    labels = [item['service_attended'] or "Not Specified" for item in qs]
    counts = [item['count'] for item in qs]

    return JsonResponse({'labels': labels, 'counts': counts})



@login_required
def channel_breakdown(request):
    """AJAX endpoint for channel of visit table."""
    
    queryset = GuestEntry.objects.all() 

    qs = queryset.values('channel_of_visit').annotate(count=Count('id')).order_by('-count')
    total = sum(item['count'] for item in qs)
    data = [
        {
            'label': item['channel_of_visit'] or 'Unknown',
            'count': item['count'],
            'percent': round((item['count'] / total) * 100, 2) if total else 0
        }
        for item in qs
    ]
    #return JsonResponse({"disabled": True})
    return JsonResponse(data, safe=False)




User = get_user_model()


from django.core.cache import cache
from django.db.models import Q, Count, Max, Prefetch
from django.core.paginator import Paginator
from django.utils.timesince import timesince
from django.utils.timezone import now, make_aware, is_naive
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.utils.http import urlencode
from datetime import datetime, timedelta

# ---------- guest_list_view ----------
@login_required
def guest_list_view(request):
    user = request.user
    role = user.username

    # ---------------------- GET FILTERS ----------------------
    search_query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', '').strip()
    channel_filter = request.GET.get('channel', '').strip()
    purpose_filter = request.GET.get('purpose', '').strip()
    service_filter = request.GET.get('service', '').strip()
    user_filter = request.GET.get('user_filter', '').strip()
    date_of_visit_filter = request.GET.get('date_of_visit', '').strip()
    view_type = request.GET.get('view', 'cards')

    # ---------------------- Base queryset ----------------------
    if is_magnet_admin(user):
        qs = GuestEntry.objects.all()
    else:
        qs = GuestEntry.objects.filter(
            Q(assigned_to=user) |
            Q(full_name="Wunmi Jordan")
        )

    # ---------------------- Annotations ----------------------
    qs = qs.annotate(
        report_count=Count('reports', distinct=True),
        last_reported=Max('reports__report_date'),
        unread_reviews=Count('reviews', filter=Q(reviews__is_read=False), distinct=True)
    )

    # ---------------------- Filters ----------------------
    filters = {
        "status__iexact": status_filter,
        "channel_of_visit__iexact": channel_filter,
        "purpose_of_visit__iexact": purpose_filter,
        "service_attended__iexact": service_filter,
        "date_of_visit": date_of_visit_filter
    }
    qs = qs.filter(**{k: v for k, v in filters.items() if v})

    # ---------------------- Search ----------------------
    if search_query:
        qs = qs.filter(
            Q(full_name__icontains=search_query) |
            Q(phone_number__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(referrer_name__icontains=search_query) |
            Q(service_attended__icontains=search_query) |
            Q(status__icontains=search_query) |
            Q(channel_of_visit__icontains=search_query) |
            Q(purpose_of_visit__icontains=search_query) |
            Q(assigned_to__full_name__icontains=search_query)
        )

    qs = qs.order_by('-custom_id').select_related('assigned_to')  # avoid n+1 on foreign key

    if user_filter:
        qs = qs.filter(assigned_to_id=user_filter)

    # ---------------------- Pagination ----------------------
    per_page = 50 if view_type == 'list' else 45
    paginator = Paginator(qs, per_page)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    # ---------------------- Field Data (lightweight) ----------------------
    excluded_fields = {"id", "custom_id", "title", "full_name", "gender",
                       "message", "picture", "phone_number", "assigned_to"}
    
    svg_icons={
        "email":"""<path stroke="none" d="M0 0h24v24H0z" fill="none"/><path d="M12 12m-4 0a4 4 0 1 0 8 0a4 4 0 1 0 -8 0" />
            <path d="M16 12v1.5a2.5 2.5 0 0 0 5 0v-1.5a9 9 0 1 0 -5.5 8.28" />""",  # replace with real paths
        "date_of_birth":"""<path stroke="none" d="M0 0h24v24H0z" fill="none"/><path d="M4 7a2 2 0 0 1 2 -2h12a2 2 0 0 1 2 2v12a2 2 0 0 1 -2 2h-12a2 2 0 0 1 -2 -2v-12z" />
            <path d="M16 3v4" /><path d="M8 3v4" />
            <path d="M4 11h16" /><path d="M11 15h1" /><path d="M12 15v3" />""",
        "age_range":"""<path stroke="none" d="M0 0h24v24H0z" fill="none"/>
            <path d="M12 12m-9 0a9 9 0 1 0 18 0a9 9 0 1 0 -18 0" />
            <path d="M11.5 10.5m-1.5 0a1.5 1.5 0 1 0 3 0a1.5 1.5 0 1 0 -3 0" />
            <path d="M11.5 13.5m-1.5 0a1.5 1.5 0 1 0 3 0a1.5 1.5 0 1 0 -3 0" />
            <path d="M7 15v-6" /><path d="M15.5 12h3" /><path d="M17 10.5v3" />""",
        "marital_status":"""<path stroke="none" d="M0 0h24v24H0z" fill="none"/><path d="M7 5m-2 0a2 2 0 1 0 4 0a2 2 0 1 0 -4 0" /><path d="M5 22v-5l-1 -1v-4a1 1 0 0 1 1 -1h4a1 1 0 0 1 1 1v4l-1 1v5" /><path d="M17 5m-2 0a2 2 0 1 0 4 0a2 2 0 1 0 -4 0" />
            <path d="M15 22v-4h-2l2 -6a1 1 0 0 1 1 -1h2a1 1 0 0 1 1 1l2 6h-2v4" />""",
        "home_address":"""<path stroke="none" d="M0 0h24v24H0z" fill="none"/><path d="M5 12l-2 0l9 -9l9 9l-2 0" /><path d="M5 12v7a2 2 0 0 0 2 2h10a2 2 0 0 0 2 -2v-7" />
            <path d="M9 21v-6a2 2 0 0 1 2 -2h2a2 2 0 0 1 2 2v6" />""",
        "occupation":"""<path stroke="none" d="M0 0h24v24H0z" fill="none"/><path d="M3 7m0 2a2 2 0 0 1 2 -2h14a2 2 0 0 1 2 2v9a2 2 0 0 1 -2 2h-14a2 2 0 0 1 -2 -2z" /><path d="M8 7v-2a2 2 0 0 1 2 -2h4a2 2 0 0 1 2 2v2" />
            <path d="M12 12l0 .01" /><path d="M3 13a20 20 0 0 0 18 0" />""",
        "date_of_visit":"""<path stroke="none" d="M0 0h24v24H0z" fill="none"/><path d="M11.795 21h-6.795a2 2 0 0 1 -2 -2v-12a2 2 0 0 1 2 -2h12a2 2 0 0 1 2 2v4" />
            <path d="M18 18m-4 0a4 4 0 1 0 8 0a4 4 0 1 0 -8 0" /><path d="M15 3v4" />
            <path d="M7 3v4" /><path d="M3 11h16" /><path d="M18 16.496v1.504l1 1" />""",
        "purpose_of_visit":"""<path stroke="none" d="M0 0h24v24H0z" fill="none"/>
            <path d="M3 21l18 0" /><path d="M10 21v-4a2 2 0 0 1 4 0v4" /><path d="M10 5l4 0" />
            <path d="M12 3l0 5" /><path d="M6 21v-7m-2 2l8 -8l8 8m-2 -2v7" />""",
        "channel_of_visit":"""<path stroke="none" d="M0 0h24v24H0z" fill="none"/><path d="M3 7m0 2a2 2 0 0 1 2 -2h14a2 2 0 0 1 2 2v9a2 2 0 0 1 -2 2h-14a2 2 0 0 1 -2 -2z" />
            <path d="M16 3l-4 4l-4 -4" />""",
        "service_attended":"""<path stroke="none" d="M0 0h24v24H0z" fill="none"/><path d="M12 5m-1 0a1 1 0 1 0 2 0a1 1 0 1 0 -2 0" />
            <path d="M7 20h8l-4 -4v-7l4 3l2 -2" />""",
        "referrer_name":"""<path stroke="none" d="M0 0h24v24H0z" fill="none"/><path d="M19.5 12.572l-7.5 7.428l-7.5 -7.428a5 5 0 1 1 7.5 -6.566a5 5 0 1 1 7.5 6.572" /><path d="M12 6l-3.293 3.293a1 1 0 0 0 0 1.414l.543 .543c.69 .69 1.81 .69 2.5 0l1 -1a3.182 3.182 0 0 1 4.5 0l2.25 2.25" />
            <path d="M12.5 15.5l2 2" /><path d="M15 13l2 2" />""",
        "referrer_phone_number":"""<path stroke="none" d="M0 0h24v24H0z" fill="none"/><path d="M3 3m0 2a2 2 0 0 1 2 -2h8a2 2 0 0 1 2 2v14a2 2 0 0 1 -2 2h-8a2 2 0 0 1 -2 -2z" />
            <path d="M8 4l2 0" /><path d="M9 17l0 .01" /><path d="M21 6l-2 3l2 3l-2 3l2 3" />""",
        "status":"""<path stroke="none" d="M0 0h24v24H0z" fill="none"/><path d="M9 11a3 3 0 1 0 6 0a3 3 0 0 0 -6 0" />
            <path d="M14.997 19.317l-1.583 1.583a2 2 0 0 1 -2.827 0l-4.244 -4.243a8 8 0 1 1 13.657 -5.584" />
            <path d="M19 22v.01" /><path d="M19 19a2.003 2.003 0 0 0 .914 -3.782a1.98 1.98 0 0 0 -2.414 .483" />""",
        "assigned_at":"""<path stroke="none" d="M0 0h24v24H0z" fill="none"/>
            <path d="M8 7a4 4 0 1 0 8 0a4 4 0 0 0 -8 0" /><path d="M16 19h6" />
            <path d="M19 16v6" /><path d="M6 21v-2a4 4 0 0 1 4 -4h4" />""",
    }

    # Precompute svg_icons and avoid per-row heavy calculations
    for guest in page_obj:
        fields = []
        for field in (f for f in guest._meta.fields if f.name not in excluded_fields):
            value = getattr(guest, field.name, None)
            display_value = getattr(guest, f"get_{field.name}_display")() if field.choices else value
            fields.append({
                "name": field.name,
                "verbose_name": getattr(field, "verbose_name", field.name).title(),
                "value": display_value,
                "icon": field.name,
                "svg": svg_icons.get(field.name)
            })
        guest.field_data = fields
        guest.has_unread_reviews = guest.unread_reviews > 0
        guest.is_new = guest.assigned_at and (now() - guest.assigned_at <= timedelta(days=14))

    # ---------------------- Cached filters ----------------------
    def cache_list(key, field):
        return cache.get_or_set(
            key,
            lambda: list(GuestEntry.objects.values_list(field, flat=True).distinct().order_by(field)),
            600
        )

    channels = cache_list("guest_channels", 'channel_of_visit')
    purposes = cache_list("guest_purposes", 'purpose_of_visit')
    services = cache_list("guest_services", 'service_attended')

    # ---------------------- Query string ----------------------
    params = request.GET.copy()
    params.pop('page', None)
    query_string = urlencode(params)

    # ---------------------- Magnet users ----------------------
    magnet_users = CustomUser.objects.filter(
        is_superuser=False,
        team_memberships__team__name__iexact="Magnet",
        is_active=True
    ).exclude(groups__name__in=['Project Admin']).distinct().order_by("full_name")

    magnet_team = Team.objects.filter(name__iexact="magnet").first()

    context = {
        'page_obj': page_obj,
        'view_type': view_type,
        'users': User.objects.filter(is_active=True).order_by('full_name')[:100],
        'search_query': search_query,
        'status_filter': status_filter,
        'channel_filter': channel_filter,
        'purpose_filter': purpose_filter,
        'service_filter': service_filter,
        'user_filter': user_filter,
        'date_of_visit': date_of_visit_filter,
        'show_filters': True,
        'statuses': [s[0] for s in GuestEntry.STATUS_CHOICES],
        'channels': channels,
        'purposes': purposes,
        'services': services,
        'query_string': query_string,
        'role': role,
        'svg_icons': svg_icons,
        'excluded_fields': excluded_fields,
        "magnet_team": {"id": magnet_team.id, "name": magnet_team.name} if magnet_team else None,
        "magnet_users": magnet_users,
        'page_title': 'Guests',
    }
    return render(request, 'guests/guest_list.html', context)






from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
#from django.utils.dateformat import format as dj_format


# ---- Chat Guest Card Detail ----
@login_required
def guest_detail_api(request, guest_id):
    guest = get_object_or_404(GuestEntry, id=guest_id)

    # Build field data just like in guest_list_view
    excluded_fields = {
        "id", "custom_id", "title", "full_name", "gender",
        "message", "picture", "phone_number", "assigned_to"
    }

    svg_icons={
        "email":"""<path stroke="none" d="M0 0h24v24H0z" fill="none"/><path d="M12 12m-4 0a4 4 0 1 0 8 0a4 4 0 1 0 -8 0" />
            <path d="M16 12v1.5a2.5 2.5 0 0 0 5 0v-1.5a9 9 0 1 0 -5.5 8.28" />""",  # replace with real paths
        "date_of_birth":"""<path stroke="none" d="M0 0h24v24H0z" fill="none"/><path d="M4 7a2 2 0 0 1 2 -2h12a2 2 0 0 1 2 2v12a2 2 0 0 1 -2 2h-12a2 2 0 0 1 -2 -2v-12z" />
            <path d="M16 3v4" /><path d="M8 3v4" />
            <path d="M4 11h16" /><path d="M11 15h1" /><path d="M12 15v3" />""",
        "age_range":"""<path stroke="none" d="M0 0h24v24H0z" fill="none"/>
            <path d="M12 12m-9 0a9 9 0 1 0 18 0a9 9 0 1 0 -18 0" />
            <path d="M11.5 10.5m-1.5 0a1.5 1.5 0 1 0 3 0a1.5 1.5 0 1 0 -3 0" />
            <path d="M11.5 13.5m-1.5 0a1.5 1.5 0 1 0 3 0a1.5 1.5 0 1 0 -3 0" />
            <path d="M7 15v-6" /><path d="M15.5 12h3" /><path d="M17 10.5v3" />""",
        "marital_status":"""<path stroke="none" d="M0 0h24v24H0z" fill="none"/><path d="M7 5m-2 0a2 2 0 1 0 4 0a2 2 0 1 0 -4 0" /><path d="M5 22v-5l-1 -1v-4a1 1 0 0 1 1 -1h4a1 1 0 0 1 1 1v4l-1 1v5" /><path d="M17 5m-2 0a2 2 0 1 0 4 0a2 2 0 1 0 -4 0" />
            <path d="M15 22v-4h-2l2 -6a1 1 0 0 1 1 -1h2a1 1 0 0 1 1 1l2 6h-2v4" />""",
        "home_address":"""<path stroke="none" d="M0 0h24v24H0z" fill="none"/><path d="M5 12l-2 0l9 -9l9 9l-2 0" /><path d="M5 12v7a2 2 0 0 0 2 2h10a2 2 0 0 0 2 -2v-7" />
            <path d="M9 21v-6a2 2 0 0 1 2 -2h2a2 2 0 0 1 2 2v6" />""",
        "occupation":"""<path stroke="none" d="M0 0h24v24H0z" fill="none"/><path d="M3 7m0 2a2 2 0 0 1 2 -2h14a2 2 0 0 1 2 2v9a2 2 0 0 1 -2 2h-14a2 2 0 0 1 -2 -2z" /><path d="M8 7v-2a2 2 0 0 1 2 -2h4a2 2 0 0 1 2 2v2" />
            <path d="M12 12l0 .01" /><path d="M3 13a20 20 0 0 0 18 0" />""",
        "date_of_visit":"""<path stroke="none" d="M0 0h24v24H0z" fill="none"/><path d="M11.795 21h-6.795a2 2 0 0 1 -2 -2v-12a2 2 0 0 1 2 -2h12a2 2 0 0 1 2 2v4" />
            <path d="M18 18m-4 0a4 4 0 1 0 8 0a4 4 0 1 0 -8 0" /><path d="M15 3v4" />
            <path d="M7 3v4" /><path d="M3 11h16" /><path d="M18 16.496v1.504l1 1" />""",
        "purpose_of_visit":"""<path stroke="none" d="M0 0h24v24H0z" fill="none"/>
            <path d="M3 21l18 0" /><path d="M10 21v-4a2 2 0 0 1 4 0v4" /><path d="M10 5l4 0" />
            <path d="M12 3l0 5" /><path d="M6 21v-7m-2 2l8 -8l8 8m-2 -2v7" />""",
        "channel_of_visit":"""<path stroke="none" d="M0 0h24v24H0z" fill="none"/><path d="M3 7m0 2a2 2 0 0 1 2 -2h14a2 2 0 0 1 2 2v9a2 2 0 0 1 -2 2h-14a2 2 0 0 1 -2 -2z" />
            <path d="M16 3l-4 4l-4 -4" />""",
        "service_attended":"""<path stroke="none" d="M0 0h24v24H0z" fill="none"/><path d="M12 5m-1 0a1 1 0 1 0 2 0a1 1 0 1 0 -2 0" />
            <path d="M7 20h8l-4 -4v-7l4 3l2 -2" />""",
        "referrer_name":"""<path stroke="none" d="M0 0h24v24H0z" fill="none"/><path d="M19.5 12.572l-7.5 7.428l-7.5 -7.428a5 5 0 1 1 7.5 -6.566a5 5 0 1 1 7.5 6.572" /><path d="M12 6l-3.293 3.293a1 1 0 0 0 0 1.414l.543 .543c.69 .69 1.81 .69 2.5 0l1 -1a3.182 3.182 0 0 1 4.5 0l2.25 2.25" />
            <path d="M12.5 15.5l2 2" /><path d="M15 13l2 2" />""",
        "referrer_phone_number":"""<path stroke="none" d="M0 0h24v24H0z" fill="none"/><path d="M3 3m0 2a2 2 0 0 1 2 -2h8a2 2 0 0 1 2 2v14a2 2 0 0 1 -2 2h-8a2 2 0 0 1 -2 -2z" />
            <path d="M8 4l2 0" /><path d="M9 17l0 .01" /><path d="M21 6l-2 3l2 3l-2 3l2 3" />""",
        "status":"""<path stroke="none" d="M0 0h24v24H0z" fill="none"/><path d="M9 11a3 3 0 1 0 6 0a3 3 0 0 0 -6 0" />
            <path d="M14.997 19.317l-1.583 1.583a2 2 0 0 1 -2.827 0l-4.244 -4.243a8 8 0 1 1 13.657 -5.584" />
            <path d="M19 22v.01" /><path d="M19 19a2.003 2.003 0 0 0 .914 -3.782a1.98 1.98 0 0 0 -2.414 .483" />""",
        "assigned_at":"""<path stroke="none" d="M0 0h24v24H0z" fill="none"/>
            <path d="M8 7a4 4 0 1 0 8 0a4 4 0 0 0 -8 0" /><path d="M16 19h6" />
            <path d="M19 16v6" /><path d="M6 21v-2a4 4 0 0 1 4 -4h4" />""",
    }

    field_data = []
    for field in guest._meta.fields:
        if field.name in excluded_fields:
            continue
        value = getattr(guest, field.name)
        time_since = None
        formatted_value = value
        if field.name == "date_of_visit" and value:
            formatted_value = value.strftime("%b. %-d, %Y")
            dt_value = datetime.combine(value, datetime.min.time())
            if is_naive(dt_value):
                dt_value = make_aware(dt_value)
            delta = timesince(dt_value, now())
            time_since = delta.split(",")[0]

        if field.name == "assigned_at" and value:
            formatted_value = value.strftime("%b. %-d, %Y %I:%M %p")

        field_data.append({
            "name": field.name,
            "verbose_name": field.verbose_name.title(),
            "value": formatted_value,
            "time_since": time_since,
            "icon": field.name,
        })

    social_accounts = []
    for account in guest.social_media_accounts.all():
        social_accounts.append({
            "platform": account.platform,
            "handle": account.handle,
        })

    data = {
        "id": guest.id,
        "custom_id": guest.custom_id,
        "title": guest.title,
        "full_name": guest.full_name,
        "phone_number": guest.phone_number,
        "picture": guest.picture.url if guest.picture else None,
        "field_data": field_data,
        "social_media_accounts": social_accounts,
        "svg_icons": svg_icons,
    }

    return JsonResponse(data)




@user_passes_test(lambda u: is_magnet_admin(u))

@login_required
def create_guest(request):

    if request.method == 'POST':
        form = GuestEntryForm(request.POST or None, request.FILES or None, user=request.user)
        social_media_types = request.POST.getlist('social_media_type[]')
        social_media_handles = request.POST.getlist('social_media_handle[]')
        social_media_entries = []
        errors = []

        # Validate social media entries
        for i, (platform, handle) in enumerate(zip(social_media_types, social_media_handles)):
            platform = platform.strip()
            handle = handle.strip()
            if platform and handle:
                if platform not in dict(SocialMediaEntry.SOCIAL_MEDIA_CHOICES):
                    errors.append(f"Invalid social media platform at entry {i+1}.")
                elif len(handle) > 255:
                    errors.append(f"Handle too long at entry {i+1}.")
                else:
                    social_media_entries.append({'platform': platform, 'handle': handle})
            elif platform or handle:
                errors.append(f"Both platform and handle must be provided at entry {i+1}.")

        if form.is_valid() and not errors:
            guest = form.save(commit=False)
            guest.save()

            # Save social media
            for entry in social_media_entries:
                SocialMediaEntry.objects.create(guest=guest, **entry)

            # Handle redirect based on button clicked
            if 'save_add_another' in request.POST:
                return redirect('create_guest')  # stay on the form
            else:  # default Save button
                return redirect('guest_list')  # regular user guest list

        # If form invalid
        return render(request, 'guests/guest_form.html', {
            'form': form,
            'social_media_errors': errors,
            'edit_mode': False
        })

    else:
        form = GuestEntryForm(user=request.user)
        return render(request, 'guests/guest_form.html', {
            'form': form,
            'edit_mode': False,
            'page_title': 'Guests',
        })




@login_required
def edit_guest(request, pk):
    guest = get_object_or_404(GuestEntry, pk=pk)
    user = request.user

    # Permissions
    if guest.full_name == "Wunmi Jordan":
        # Allow everyone to edit this guest, but restrict certain actions
        pass
    elif not (is_magnet_admin(user) or guest.assigned_to == user):
        messages.error(request, "You do not have permission to edit this guest.")
        return redirect('guest_list')

    reassign_allowed = is_magnet_admin(request.user)
    all_users = User.objects.filter(is_active=True).order_by('full_name') if reassign_allowed else None
    social_media_entries = guest.social_media_accounts.all()

    if request.method == "POST":
        if "delete_guest" in request.POST:
            # Restrict deleting "Wunmi Jordan"
            if guest.full_name == "Wunmi Jordan" and not request.user.is_superuser:
                messages.error(request, "Only superusers can delete this guest.")
                return redirect("guest_list")

            guest.delete()
            messages.success(request, f"{guest.full_name} was deleted successfully.")
            return redirect("guest_list")

        form = GuestEntryForm(request.POST, request.FILES, instance=guest)

        # 🔒 Lock the name field in backend
        if guest.full_name == "Wunmi Jordan":
            form.fields["full_name"].disabled = True  # Prevent UI editing

        social_media_types = request.POST.getlist('social_media_type[]')
        social_media_handles = request.POST.getlist('social_media_handle[]')
        social_media_data = []
        errors = []

        for i, (platform, handle) in enumerate(zip(social_media_types, social_media_handles)):
            platform = platform.strip()
            handle = handle.strip()
            if platform and handle:
                if platform not in dict(SocialMediaEntry.SOCIAL_MEDIA_CHOICES):
                    errors.append(f"Invalid social media platform at entry {i+1}.")
                elif len(handle) > 255:
                    errors.append(f"Handle too long at entry {i+1}.")
                else:
                    social_media_data.append({'platform': platform, 'handle': handle})
            elif platform or handle:
                errors.append(f"Both platform and handle must be provided at entry {i+1}.")

        if form.is_valid() and not errors:
            updated_guest = form.save(commit=False)

            # 🔒 Ensure full_name remains unchanged
            if guest.full_name == "Wunmi Jordan":
                updated_guest.full_name = guest.full_name

            # Handle reassignment
            if reassign_allowed and 'assigned_to' in request.POST:
                assigned_id = request.POST.get('assigned_to')
                updated_guest.assigned_to = User.objects.filter(pk=assigned_id).first() if assigned_id else None

            # Clear picture if requested
            if 'clear_picture' in request.POST and guest.picture:
                guest.picture.delete(save=False)
                updated_guest.picture = None

            updated_guest.save()

            # Replace social media entries
            guest.social_media_accounts.all().delete()
            for entry in social_media_data:
                SocialMediaEntry.objects.create(guest=guest, **entry)

            # Redirect based on button clicked
            if 'save_add_another' in request.POST:
                return redirect('create_guest')  # Stay on new guest form
            else:  # default Save button
                return redirect('guest_list')

    else:
        form = GuestEntryForm(instance=guest, user=request.user)

        # 🔒 Lock the field in UI (read-only)
        if guest.full_name == "Wunmi Jordan":
            form.fields["full_name"].disabled = True

    return render(request, 'guests/guest_form.html', {
        'form': form,
        'guest': guest,
        'edit_mode': True,
        'can_reassign': reassign_allowed,
        'all_users': all_users,
        'show_delete': True,
        'social_media_entries': social_media_entries,
        'page_title': 'Guests',
    })




from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from .models import Review
from notifications.utils import user_full_name

@login_required
@require_POST
def submit_review(request, guest_id, role):
    """
    Submit a review for a guest.
    Signals handle notifications automatically.
    Redirects back to guest list.
    """
    guest = get_object_or_404(GuestEntry, id=guest_id)
    parent_id = request.POST.get("parent_id")
    parent = Review.objects.filter(id=parent_id).first() if parent_id else None

    Review.objects.create(
        guest=guest,
        reviewer=request.user,
        role=role,
        comment=request.POST.get("comment"),
        parent=parent
    )

    return redirect("guest_list")





from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from .models import Review
from guests.models import GuestEntry
from notifications.models import Notification

@login_required
def mark_reviews_read(request, guest_id):
    """
    Mark all unread reviews for the current user on a given guest as read,
    and also mark the related notifications as read.
    Returns JSON with counts for updating UI dynamically.
    """
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Invalid request method"}, status=400)

    guest = get_object_or_404(GuestEntry, id=guest_id)

    # Only unread reviews for this user, optionally excluding self-authored reviews
    unread_reviews = guest.reviews.filter(is_read=False).exclude(reviewer=request.user)
    reviews_marked = unread_reviews.update(is_read=True)

    # Mark notifications corresponding to these reviews as read
    notif_qs = Notification.objects.filter(
        user=request.user,
        link__icontains=f"guest/{guest_id}/review",  # adjust to match your review URL pattern
        is_read=False
    )
    notifs_marked = notif_qs.update(is_read=True)

    return JsonResponse({
        "status": "success",
        "reviews_marked": reviews_marked,
        "notifications_marked": notifs_marked
    })






def superuser_required(view_func):
    return user_passes_test(lambda u: u.is_superuser)(view_func)

@login_required
@superuser_required
def bulk_delete_guests(request):
    if request.method == "POST" and request.headers.get("x-requested-with") == "XMLHttpRequest":
        guest_ids = request.POST.getlist("guest_ids[]")
        if not guest_ids:
            return JsonResponse({"success": False, "message": "No guests selected."})
        
        deleted_count, _ = GuestEntry.objects.filter(id__in=guest_ids).delete()
        return JsonResponse({"success": True, "deleted_count": deleted_count})

    return JsonResponse({"success": False, "message": "Invalid request."})




# -------------------------
# Reassign guest
# -------------------------
@login_required
@user_passes_test(lambda u: is_magnet_admin(u))
def reassign_guest(request, guest_id):
    guest = get_object_or_404(GuestEntry, id=guest_id)

    if request.method == 'POST':
        assigned_to_id = request.POST.get('assigned_to')

        if assigned_to_id:
            assigned_user = User.objects.filter(id=assigned_to_id, is_active=True).first()
            if assigned_user:
                guest.assigned_to = assigned_user
                guest.save()
                messages.success(request, f"Guest {guest.full_name} reassigned to {assigned_user.get_full_name() or assigned_user.username}.")
            else:
                messages.error(request, "Selected user does not exist or is inactive.")
        else:
            guest.assigned_to = None
            guest.save()
            messages.success(request, f"Assignment cleared for guest {guest.full_name}.")

    # Redirect admins back to admin dashboard
    return redirect('guest_list')






@require_POST
@login_required
def update_guest_status(request, pk):
    """
    Updates a guest's follow-up status (via dropdown in guest_list).
    Only the creator or an admin can update.
    """
    guest = get_object_or_404(GuestEntry, pk=pk)
    user = request.user

    if not (is_magnet_admin(user) or guest.assigned_to == user):
        return redirect('guest_list')

    new_status = request.POST.get('status')
    if new_status in dict(GuestEntry.STATUS_CHOICES):
        guest.status = new_status
        guest.save()

    return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))



def parse_flexible_date(date_str):
    """
    Try multiple date formats and return a valid `date` object or None.
    """
    date_formats = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"]
    for fmt in date_formats:
        try:
            return datetime.strptime(date_str.strip(), fmt).date()
        except (ValueError, AttributeError):
            continue
    return None



User = get_user_model()

@login_required
def import_guests_csv(request):
    if request.method != "POST" or not request.FILES.get("csv_file"):
        messages.error(request, "Please upload a valid CSV file.")
        return redirect("guest_list")

    csv_file = request.FILES["csv_file"]
    decoded_file = csv_file.read().decode("utf-8").splitlines()
    reader = csv.DictReader(decoded_file)

    guests_to_create = []

    # --- Step 1: Prepare GuestEntry objects ---
    for row in reader:
        username = row.get("assigned_to", "").strip()
        try:
            user = User.objects.get(username=username)
        except User.DoesNotExist:
            messages.warning(request, f"User '{username}' not found. Skipping row.")
            continue

        dob = row.get("date_of_birth", "").strip() or None
        dov = row.get("date_of_visit", "").strip() or None

        guest = GuestEntry(
            full_name=row.get("full_name", "").strip(),
            title=row.get("title", "").strip(),
            gender=row.get("gender", "").strip(),
            phone_number=row.get("phone_number", "").strip(),
            email=row.get("email", "").strip(),
            date_of_birth=dob,
            marital_status=row.get("marital_status", "").strip(),
            home_address=row.get("home_address", "").strip(),
            occupation=row.get("occupation", "").strip(),
            date_of_visit=dov,
            purpose_of_visit=row.get("purpose_of_visit", "").strip(),
            channel_of_visit=row.get("channel_of_visit", "").strip(),
            service_attended=row.get("service_attended", "").strip(),
            referrer_name=row.get("referrer_name", "").strip(),
            referrer_phone_number=row.get("referrer_phone_number", "").strip(),
            message=row.get("message", "").strip(),
            status=row.get("status", "").strip(),
            assigned_to=user,
            picture=row.get("picture_url", "").strip() or None  # <-- store Cloudinary URL directly
        )
        guests_to_create.append(guest)

    if not guests_to_create:
        messages.warning(request, "No valid guests found to import.")
        return redirect("guest_list")

    # --- Step 2: Bulk create guests ---
    with transaction.atomic():
        GuestEntry.objects.bulk_create(guests_to_create)

        # --- Step 3: Backfill custom_id ---
        prefix = "GNG"
        last_custom_id = GuestEntry.objects.filter(custom_id__startswith=prefix)\
            .aggregate(max_id=Max('custom_id'))['max_id']
        last_num = int(last_custom_id.replace(prefix, "")) if last_custom_id else 0

        new_guests = GuestEntry.objects.filter(custom_id__isnull=True).order_by("id")
        for idx, guest in enumerate(new_guests, start=last_num + 1):
            guest.custom_id = f"{prefix}{idx:06d}"

        GuestEntry.objects.bulk_update(new_guests, ["custom_id"])

    messages.success(request, f"{len(guests_to_create)} guests imported successfully!")
    return redirect("guest_list")






def download_csv_template(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="guest_import_template.csv"'

    writer = csv.writer(response)

    # Write header row
    writer.writerow([
        'full_name',               # Required
        'title',                   # Optional (Mr, Mrs, Miss, etc.)
        'gender',                  # Optional (Male, Female, Other)
        'phone_number',            # Optional
        'email',                   # Optional
        'date_of_birth',           # Optional (Any format like YYYY-MM-DD, DD/MM/YYYY, etc.)
        'marital_status',          # Optional
        'home_address',            # Optional
        'occupation',              # Optional
        'date_of_visit',           # Optional (Any format like YYYY-MM-DD, 24 July 2025, etc.)
        'purpose_of_visit',        # Optional
        'channel_of_visit',        # Optional (Flyer, Friend, Social Media, etc.)
        'service_attended',        # Optional (Sunday, Midweek, etc.)
        'referrer_name',           # Optional
        'referrer_phone_number',   # Optional
        'message',                 # Optional
        'status',                  # Optional (New, Returned, Not Interested, etc.)
        'assigned_to',              # Required (must match an existing username)
    ])

    # Optionally include one empty sample row
    writer.writerow([
        '', '', '', '', '',
        '', '', '', '',
        '', '', '', '',
        '', '', '', '', ''
    ])

    return response





@login_required
def export_csv(request):
    """
    Export filtered guest entries as CSV.
    - Admins can export all or filter by user.
    - Regular users can only export their own entries.
    - Respects service and search filters.
    """
    User = get_user_model()

    filter_user_id = request.GET.get('user')
    filter_service = request.GET.get('service')
    search_query = request.GET.get('q')

    # Base queryset
    if is_magnet_admin(request.user):
        guests = GuestEntry.objects.all()
        if filter_user_id and filter_user_id.isdigit():
            guests = guests.filter(assigned_to__id=filter_user_id)
    else:
        guests = GuestEntry.objects.filter(assigned_to=request.user)

    # Service filter
    if filter_service:
        guests = guests.filter(service_attended__iexact=filter_service)

    # Search filter
    if search_query:
        guests = guests.filter(
            Q(full_name__icontains=search_query) |
            Q(phone_number__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(referrer_name__icontains=search_query)
        )

    # CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="guest_entries.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'Full Name', 'Phone Number', 'Email', 'Gender',
        'Date of Birth', 'Marital Status', 'Home Address',
        'Occupation', 'Date of Visit', 'Purpose of Visit',
        'Channel of Visit', 'Service Attended',
        'Referrer Name', 'Referrer Phone Number', 'Status',
        'Assigned To'
    ])

    for guest in guests:
        writer.writerow([
            guest.full_name,
            guest.phone_number,
            guest.email,
            guest.gender,
            guest.date_of_birth,
            guest.marital_status,
            guest.home_address,
            guest.occupation,
            guest.date_of_visit,
            guest.purpose_of_visit,
            guest.channel_of_visit,
            guest.service_attended,
            guest.referrer_name,
            guest.referrer_phone_number,
            guest.status,
            guest.assigned_to.get_full_name() if guest.assigned_to else '',
        ])

    return response


@login_required
def update_status_view(request, guest_id, status_key):
    guest = get_object_or_404(GuestEntry, id=guest_id)

    # Only allow if user is the creator or admin
    if request.user != guest.assigned_to and not request.user.is_superuser:
        return redirect('guest_list')  # or return an HTTP 403 Forbidden response

    guest.status = status_key
    guest.save()
    return redirect('guest_list')


from datetime import datetime

def safe_date(value):
    if not value:
        return ""
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value).strftime("%Y-%m-%d")
        except:
            # Try flexible formats
            for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
                try:
                    return datetime.strptime(value, fmt).strftime("%Y-%m-%d")
                except:
                    pass
            return value  # return raw string if all parsing fails
    return value.strftime("%Y-%m-%d")





@login_required
def export_guests_excel(request):
    filter_user_id = request.GET.get('user')
    filter_service = request.GET.get('service')
    search_query = request.GET.get('q')

    guests = GuestEntry.objects.all() if is_magnet_admin(request.user) else GuestEntry.objects.filter(assigned_to=request.user)

    if filter_user_id and filter_user_id.isdigit():
        guests = guests.filter(assigned_to__id=filter_user_id)

    if filter_service:
        guests = guests.filter(service_attended__iexact=filter_service)

    if search_query:
        guests = guests.filter(
            Q(full_name__icontains=search_query) |
            Q(phone_number__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(referrer_name__icontains=search_query)
        )

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Guest Entries"

    headers = [
        'Full Name', 'Phone Number', 'Email', 'Gender',
        'Date of Birth', 'Marital Status', 'Home Address',
        'Occupation', 'Date of Visit', 'Purpose of Visit',
        'Channel of Visit', 'Service Attended',
        'Referrer Name', 'Referrer Phone Number', 'Status',
        'Assigned To'
    ]
    ws.append(headers)

    for guest in guests:
        ws.append([
            f"{guest.title} {guest.full_name}",
            guest.phone_number,
            guest.email,
            guest.gender,
            safe_date(guest.date_of_birth),
            guest.marital_status,
            guest.home_address,
            guest.occupation,
            safe_date(guest.date_of_visit),
            guest.purpose_of_visit,
            guest.channel_of_visit,
            guest.service_attended,
            guest.referrer_name,
            guest.referrer_phone_number,
            guest.status,
            guest.assigned_to.get_full_name() if guest.assigned_to else '',
        ])

    # ---- ASGI-SAFE: Write to BytesIO first ----
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=guest_entries.xlsx'
    return response






#@login_required
#def export_guests_pdf(request):
#    guests = GuestEntry.objects.all()
#    html = render_to_string('guests/guest_list_pdf.html', {'guests': guests})
#    pdf_file = weasyprint.HTML(string=html).write_pdf()
#    response = HttpResponse(pdf_file, content_type='application/pdf')
#    response['Content-Disposition'] = 'filename="all_guests.pdf"'
#    return response




@login_required
@user_passes_test(lambda u: is_magnet_admin(u))  # Only staff users can import
def import_guests_excel(request):
    if request.method == "POST":
        file = request.FILES.get("excel_file")
        if not file.name.endswith(".xlsx"):
            messages.error(request, "Only .xlsx files are supported.")
            return redirect("import_guests")

        wb = openpyxl.load_workbook(file)
        sheet = wb.active

        headers = [cell.value for cell in sheet[1]]
        username_col = headers.index("Assigned To (Username)")

        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                assigned_to_username = row[username_col]
                assigned_to_user = User.objects.get(username=assigned_to_username)

                GuestEntry.objects.create(
                    title=row[0],
                    full_name=row[1],
                    gender=row[2],
                    phone_number=row[3],
                    email=row[4],
                    date_of_birth=row[5],
                    marital_status=row[6],
                    home_address=row[7],
                    occupation=row[8],
                    date_of_visit=row[9],
                    purpose_of_visit=row[10],
                    channel_of_visit=row[11],
                    service_attended=row[12],
                    referrer_name=row[13],
                    referrer_phone_number=row[14],
                    message=row[15],
                    assigned_to=assigned_to_user,
                )

            except User.DoesNotExist:
                messages.warning(request, f"User '{assigned_to_username}' not found. Skipping row.")
                continue
            except Exception as e:
                messages.error(request, f"Error importing row: {e}")
                continue

        messages.success(request, "Guests imported successfully.")
        return redirect("guest_list")

    return render(request, "guests/import_excel.html")


def get_week_start_end(target_date):
    start = target_date - timedelta(days=target_date.weekday())
    end = start + timedelta(days=6)
    return start, end





@login_required
def followup_report_page(request, guest_id):
    guest = get_object_or_404(GuestEntry, id=guest_id)
    today = localdate()
    user = request.user

    # Permission check
    if guest.full_name != "Wunmi Jordan" and not (user.is_superuser or guest.assigned_to == user):
        messages.error(request, "You do not have permission to edit this guest.")
        return redirect('guest_list')

    # Fetch reports with pagination
    reports = FollowUpReport.objects.filter(guest=guest).order_by('-report_date')
    paginator = Paginator(reports, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    report_to_edit = None
    edit_report_id = request.GET.get('edit_report_id')
    if edit_report_id:
        report_to_edit = get_object_or_404(FollowUpReport, id=edit_report_id, guest=guest)

    # Build form
    if request.method == 'POST' and 'submit_report' in request.POST:
        if report_to_edit:
            form = FollowUpReportForm(request.POST, instance=report_to_edit, guest=guest)
        else:
            form = FollowUpReportForm(request.POST, guest=guest)

        if form.is_valid():
            form.save()
            messages.success(request, "Report updated successfully." if report_to_edit else "Follow-up report created successfully.")
            return redirect('followup_report_page', guest_id=guest.id)
    else:
        if report_to_edit:
            form = FollowUpReportForm(instance=report_to_edit, guest=guest)
        else:
            form = FollowUpReportForm(guest=guest)

    return render(request, 'guests/followup_report_page.html', {
        'guest': guest,
        'reports': reports,
        'page_obj': page_obj,
        'today': today,
        'report_to_edit': report_to_edit,
        'form': form,
    })






"""
def get_guest_reports(request, guest_id):
    guest = get_object_or_404(GuestEntry, id=guest_id)
    reports = FollowUpReport.objects.filter(guest=guest).order_by('-report_date')
    report_data = [
        {
            'report_date': localtime(report.report_date).strftime('%Y-%m-%d'),
            'note': report.note,
            'sunday_attended': report.sunday_attended,
            'midweek_attended': report.midweek_attended,
        }
        for report in reports
    ]
    return JsonResponse({'reports': report_data})
"""


@login_required
def followup_history_view(request, guest_id):
    guest = get_object_or_404(GuestEntry, id=guest_id)
    reports = FollowUpReport.objects.filter(guest=guest).order_by('-report_date')
    return render(request, 'guests/followup_history.html', {
        'guest': guest,
        'reports': reports,
    })


def export_followup_reports_pdf(request, guest_id):
    guest = get_object_or_404(GuestEntry, id=guest_id)
    reports = FollowUpReport.objects.filter(guest=guest).order_by('-report_date')

    # Create a PDF buffer
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, title="Guest Reports")

    styles = getSampleStyleSheet()
    elements = []

    # Custom title style
    title_style = ParagraphStyle(
        name='Title',
        fontSize=16,
        leading=24,
        alignment=1,  # Center
        spaceAfter=20,
    )

    # Optional Logo
    logo_path = os.path.join('static', 'your_logo.png')  # Adjust path
    if os.path.exists(logo_path):
        logo = Image(logo_path, width=100, height=40)
        logo.hAlign = 'LEFT'
        elements.append(logo)

    # Title
    elements.append(Paragraph(f"Follow-Up Report for {guest.full_name}", title_style))
    elements.append(Spacer(1, 10))

    # Table header and data
    data = [['Date', 'Sunday', 'Midweek', 'Message', 'Assigned To']]
    for report in reports:
        assigned_to_name = report.assigned_to.get_full_name() if report.assigned_to else 'Unknown'
        data.append([
            report.report_date.strftime("%Y-%m-%d"),
            '✔️' if report.service_sunday else '',
            '✔️' if report.service_midweek else '',
            report.note or ''
        ])

    table = Table(data, colWidths=[80, 60, 60, 280, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#f0f0f0")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor("#000000")),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))

    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf', headers={
        'Content-Disposition': f'attachment; filename="followup_reports_{guest.id}.pdf"'
    })



from datetime import datetime, timedelta
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponseRedirect
from django.shortcuts import render
from django.utils import timezone
from django.core.exceptions import ValidationError
from workforce.models import Event, AttendanceRecord, UserActivity
from workforce.utils import validate_church_proximity
from workforce.broadcast import broadcast_attendance_summary

@login_required
def mark_attendance(request):
    """Handles attendance marking — supports AJAX for no page reload."""
    today = timezone.localdate()
    now = timezone.localtime()
    weekday = today.strftime("%A").lower()

    # ✅ Use team-aware helper to get only events user can mark
    events = get_available_events_for_user(request.user)
    available_events = []

    for e in events:
        if e.event_type == "followup":
            available_events.append(e)
        elif e.day_of_week and e.day_of_week.lower() == weekday:
            available_events.append(e)
        elif e.event_type == "meeting" and weekday == "wednesday":
            available_events.append(e)

    # Only show “other” when there are real events
    actual_events_today = [e for e in available_events if e.event_type != "followup"]
    show_other_option = len(actual_events_today) > 0

    # Weekly guest activity check
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    guest_activity_types = ["followup", "message", "guest_view", "call", "report", "other"]

    user_guest_activity = UserActivity.objects.filter(
        user=request.user,
        activity_type__in=guest_activity_types,
        created_at__date__gte=week_start,
        created_at__date__lte=week_end,
    ).exists()

    # Whether user can mark now
    can_mark_now = any(
        e.time is None or now >= timezone.make_aware(datetime.combine(today, e.time))
        for e in actual_events_today
    )

    # --- Handle POST ---
    if request.method == "POST":
        is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"

        event_id = request.POST.get("event_id", "").strip()
        remarks = request.POST.get("remarks", "").strip()
        status = request.POST.get("status", "present")
        user_lat = request.POST.get("latitude")
        user_lon = request.POST.get("longitude")
        next_url = request.POST.get("next") or request.META.get("HTTP_REFERER") or "/"

        # 🔹 Validate event_id before querying
        if not event_id:
            if is_ajax:
                return JsonResponse({"success": False, "error": "Missing event ID"}, status=400)
            messages.error(request, "Invalid or missing event ID.")
            return HttpResponseRedirect(next_url)

        if event_id == "other":
            custom_name = request.POST.get("custom_event") or "Custom Event"
            event, _ = Event.objects.get_or_create(
                name=custom_name, event_type="custom", date=today
            )
        else:
            try:
                event = Event.objects.get(id=int(event_id))
            except (Event.DoesNotExist, ValueError):
                if is_ajax:
                    return JsonResponse({"success": False, "error": "Event not found"}, status=404)
                messages.error(request, "Event not found.")
                return HttpResponseRedirect(next_url)

        # ✅ Enforce physical presence validation for physical events
        mode = (getattr(event, "attendance_mode", "") or "").lower()
        is_physical = mode == "physical"

        if status == "present" and is_physical:
            try:
                print(f"[DEBUG] Checking proximity for {request.user} at {user_lat},{user_lon}")
                validate_church_proximity(user_lat, user_lon)
            except ValidationError as e:
                print(f"[DEBUG] Validation failed: {e}")
                if is_ajax:
                    # JSON response → frontend toast, modal stays open
                    return JsonResponse({"success": False, "error": str(e)}, status=400)
                else:
                    # Fallback for non-AJAX form
                    messages.error(request, str(e))
                    return HttpResponseRedirect(next_url)

        # Determine lateness
        if status == "present" and event.time:
            event_dt = datetime.combine(today, event.time)
            if now > timezone.make_aware(event_dt + timedelta(minutes=15)):
                status = "late"

        # ✅ Check if attendance already marked today
        existing = AttendanceRecord.objects.filter(
            user=request.user,
            event=event,
            date=today
        ).first()

        if existing:
            if is_ajax:
                return JsonResponse({
                    "success": False,
                    "error": f"You have already marked attendance for {event.name} today."
                }, status=400)
            else:
                messages.warning(request, f"You have already marked attendance for {event.name} today.")
                return HttpResponseRedirect(next_url)

        # Save record
        AttendanceRecord.objects.create(
            user=request.user,
            event=event,
            date=today,
            status=status,
            remarks=remarks,
            team=event.team,
        )

        # 🔔 Notify live dashboard
        broadcast_attendance_summary()

        message_text = f"Attendance marked for {event.name} ({status})."
        if is_ajax:
            return JsonResponse({"success": True, "message": message_text})
        else:
            messages.success(request, message_text)
            return HttpResponseRedirect(next_url)

    context = {
        "events": available_events,
        "show_other_option": show_other_option,
        "skip_attendance": user_guest_activity,
        "can_mark_now": can_mark_now,
        "show_weekly_summary": today.weekday() == 5 and now.hour >= 20,
    }
    return render(request, "workforce/mark_attendance.html", context)



from datetime import timedelta
from django.utils import timezone
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from workforce.models import Event, AttendanceRecord
from django.db.models import Q

@login_required
def recent_event(request):
    """Return the most recent event (within 45 mins) if active and not yet marked by user."""
    now = timezone.localtime()
    weekday = now.strftime("%A").lower()

    # ✅ Only include events the user can attend (general + team events)
    events_today = [
        e for e in get_available_events_for_user(request.user)
        if e.date == now.date() or (e.is_recurring_weekly and e.day_of_week.lower() == weekday)
    ]

    recent_event = None
    for event in events_today:
        # If event has explicit date & time
        if event.date and event.time:
            event_dt = timezone.make_aware(
                timezone.datetime.combine(event.date, event.time),
                timezone.get_current_timezone()
            )
        # Otherwise, treat recurring events as happening "today" at their time
        elif event.day_of_week and event.time:
            event_dt = timezone.make_aware(
                timezone.datetime.combine(now.date(), event.time),
                timezone.get_current_timezone()
            )
        else:
            # Skip undated + timeless events (like follow-ups)
            continue

        # Only within 15 minutes *before* start
        window_start = event_dt - timedelta(minutes=15)
        window_end = event_dt  # exactly at start time

        print(f"Now={now}, Event={event_dt}, Match={window_start <= now <= window_end}")

        if window_start <= now <= window_end:
            recent_event = event
            break

    if not recent_event:
        return JsonResponse({"event": None})

    MARKED_STATUSES = ("present", "late", "excused")

    already_marked = AttendanceRecord.objects.filter(
        user=request.user,
        event=recent_event,
        date=now.date(),
        status__in=MARKED_STATUSES
    ).exists()

    if already_marked:
        print(f"User {request.user} already marked for event {recent_event.name}")
        return JsonResponse({"event": None})

    return JsonResponse({
        "event": {
            "id": recent_event.id,
            "name": recent_event.name,
            "date": str(now.date()),
            "time": recent_event.time.strftime("%H:%M:%S") if recent_event.time else None,
        }
    })



@login_required
def add_personal_reminder(request):
    if request.method == "POST":
        title = request.POST.get("title")
        description = request.POST.get("description")
        date = request.POST.get("date")
        time = request.POST.get("time") or None

        PersonalReminder.objects.create(
            user=request.user,
            title=title,
            description=description,
            date=date,
            time=time
        )
        messages.success(request, "Reminder added!")
        return redirect("dashboard")

    return render(request, "workforce/add_reminder.html")


from datetime import date, datetime, timedelta, time
from django.utils import timezone
from django.http import JsonResponse


from datetime import date, timedelta
from django.http import JsonResponse

def api_events(request):
    """
    Unified API endpoint:
      - FullCalendar (with start/end params)
      - Team Modal (with team_id)
    """
    user = request.user
    team_id = request.GET.get("team_id")
    start = request.GET.get("start")
    end = request.GET.get("end")

    events = get_available_events_for_user(user)
    today = timezone.localdate()

    # 🟦 CASE 1 — Modal events by team_id
    if team_id is not None:
        data = expand_team_events(user, team_id)
        return JsonResponse({"events": data})

    # 🟩 CASE 2 — FullCalendar events (default path)
    weekday_map = {
        "monday": 0, "tuesday": 1, "wednesday": 2,
        "thursday": 3, "friday": 4, "saturday": 5, "sunday": 6,
    }

    color_map = {
        "service": "#3b82f6",
        "meeting": "#10b981",
        "training": "#8b5cf6",
        "reminder": "#f59e0b",
        "event": "#e11d48",
        "other": "#9ca3af",
    }

    today = date.today()
    start_of_year = date(today.year, 1, 1)
    end_of_year = date(today.year, 12, 31)

    data = []

    def build_datetime(d, t):
        if not t:
            return d.isoformat()
        return datetime.combine(d, t).isoformat()

    for e in events:
        event_type = (e.event_type or "other").lower()
        color = color_map.get(event_type, "#4dabf7")

        base_event = {
            "title": e.name,
            "color": color,
            "extendedProps": {
                "type": e.event_type,
                "mode": getattr(e, "attendance_mode", ""),
                "location": getattr(e, "location", ""),
                "description": getattr(e, "description", ""),
                "time": e.time.strftime("%H:%M") if e.time else None,
                "team": getattr(e.team, "name", None),
                "team_color": getattr(e.team, "color_class", "bg-gray-800")
                if getattr(e, "team", None)
                else None,
            },
        }

        # Handle fixed-date events
        if e.date:
            start_date = e.date
            end_date = e.end_date or (
                start_date + timedelta(days=getattr(e, "duration_days", 1) - 1)
            )
            base_event["start"] = build_datetime(start_date, e.time)
            base_event["end"] = build_datetime(end_date, e.time)
            data.append(base_event)

        # Handle recurring events
        elif e.is_recurring_weekly and e.day_of_week:
            weekday = weekday_map.get(e.day_of_week.lower())
            if weekday is not None:
                current = start_of_year
                while current <= end_of_year:
                    if current.weekday() == weekday:
                        recurring = base_event.copy()
                        recurring["start"] = build_datetime(current, e.time)
                        data.append(recurring)
                    current += timedelta(days=1)

    return JsonResponse(data, safe=False)




from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.utils import timezone
import json
from workforce.models import UserActivity

@csrf_exempt
@login_required
def log_user_activity(request):
    """Silent logging endpoint for guest/follow-up interactions."""
    if request.method == "POST":
        try:
            data = json.loads(request.body.decode("utf-8"))
            activity_type = data.get("activity_type", "other")
            guest_id = data.get("guest_id")
            description = data.get("description", "")

            UserActivity.objects.create(
                user=request.user,
                activity_type=activity_type,
                guest_id=guest_id,
                description=description,
                created_at=timezone.now(),
            )

            return JsonResponse({"status": "success"}, status=201)
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)
    return JsonResponse({"status": "invalid_request"}, status=405)


from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from workforce.models import Event

@login_required
def get_active_events(request):
    """
    Returns active events — both one-time (with date) and recurring (without date).
    """
    events = get_available_events_for_user(request.user)
    data = []

    for e in events:
        # Handle one-time events
        if e.date:
            date_str = e.date.strftime("%Y-%m-%d")
            time_str = e.time.strftime("%H:%M") if e.time else None
            label = e.name
        else:
            # Recurring events (like weekly services)
            date_str = None
            time_str = e.time.strftime("%H:%M") if e.time else None
            label = f"{e.name} (Recurring)"

        data.append({
            "id": e.id,
            "name": label,
            "date": date_str,
            "time": time_str,
            "event_type": e.event_type,
            "is_recurring": not bool(e.date),
        })

    return JsonResponse(data, safe=False)